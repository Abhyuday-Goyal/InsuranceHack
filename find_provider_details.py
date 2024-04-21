import openpyxl

def search_in_excel(file_path, sheet_name, search_column_index, search_string, extract_columns, state=None):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the specific sheet by name
    sheet = wb[sheet_name]
    
    # Initialize variables to store matching row and values in other columns
    matching_row = None
    other_column_values = {}
    
    # Iterate over rows in the selected column
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=search_column_index, max_col=search_column_index):
        cell_value = row[0].value
        cell_state = sheet.cell(row=row[0].row, column=3).value  # Get the state from column 3
        if cell_value and search_string.lower() in str(cell_value).lower() and (state is None or cell_state == state):
            matching_row = row[0].row
            # Store values in other specified columns for the first matching row
            for col_idx in extract_columns:
                other_column_values[col_idx] = sheet.cell(row=matching_row, column=col_idx).value
            break  # Stop searching after the first matching row
    
    # Print or process matching row and values in other specified columns
    if matching_row is not None:
        print("First matching row found in column {} for search string '{}' and state '{}':".format(search_column_index, search_string, state))
        print("Row:", matching_row)
        for col_idx, value in other_column_values.items():
            print("Value in column {}: {}".format(col_idx, value))
    else:
        print("No matching rows found in column {} for search string '{}' and state '{}'".format(search_column_index, search_string, state))